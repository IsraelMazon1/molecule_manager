"""Excel export service — generate multi-sheet .xlsx workbooks for lab data."""

from __future__ import annotations

import io
import logging
import uuid
from datetime import date

import cairosvg
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.experiment import Experiment
from app.models.experiment_molecule import ExperimentMolecule
from app.models.lab import Lab
from app.models.lab_member import LabMember
from app.models.molecule import Molecule
from app.models.user import User

logger = logging.getLogger(__name__)

IMAGE_THRESHOLD = 500
HEADER_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
IMG_HEIGHT_PX = 80
IMG_WIDTH_PX = 80
ROW_HEIGHT_PT = 60  # ~80px


def _style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT
    ws.freeze_panes = "A2"


def _auto_width(ws, col_count: int, max_width: int = 40) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        best = 10
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > best:
                        best = length
        ws.column_dimensions[letter].width = min(best + 2, max_width)


def _svg_to_png(svg_text: str) -> bytes | None:
    if not svg_text:
        return None
    try:
        return cairosvg.svg2png(
            bytestring=svg_text.encode("utf-8"),
            output_width=IMG_WIDTH_PX * 2,
            output_height=IMG_HEIGHT_PX * 2,
        )
    except Exception:
        logger.debug("SVG→PNG conversion failed", exc_info=True)
        return None


def _user_email_map(db: Session, user_ids: set[uuid.UUID]) -> dict[uuid.UUID, str]:
    if not user_ids:
        return {}
    rows = db.execute(
        select(User.id, User.email).where(User.id.in_(user_ids))
    ).all()
    return {uid: email for uid, email in rows}


def export_lab_workbook(lab_id: uuid.UUID, db: Session) -> bytes:
    lab = db.get(Lab, lab_id)

    # ── Fetch all data (scoped to lab) ────────────────────────────────────────
    molecules = list(
        db.execute(
            select(Molecule).where(Molecule.lab_id == lab_id).order_by(Molecule.created_at.desc())
        ).scalars().all()
    )

    experiments = list(
        db.execute(
            select(Experiment).where(Experiment.lab_id == lab_id).order_by(Experiment.created_at.desc())
        ).scalars().all()
    )

    exp_mol_rows = list(
        db.execute(
            select(ExperimentMolecule)
            .join(Experiment, ExperimentMolecule.experiment_id == Experiment.id)
            .where(Experiment.lab_id == lab_id)
        ).scalars().all()
    )

    member_count = db.execute(
        select(func.count()).where(LabMember.lab_id == lab_id)
    ).scalar_one()

    # Build user email lookup
    user_ids: set[uuid.UUID] = set()
    for mol in molecules:
        if mol.created_by_user_id:
            user_ids.add(mol.created_by_user_id)
    for exp in experiments:
        if exp.created_by_user_id:
            user_ids.add(exp.created_by_user_id)
    emails = _user_email_map(db, user_ids)

    # Build experiment/molecule lookup maps for the links sheet
    exp_map = {e.id: e for e in experiments}
    mol_map = {m.id: m for m in molecules}

    # Experiment molecule counts
    exp_mol_counts: dict[uuid.UUID, int] = {}
    for em in exp_mol_rows:
        exp_mol_counts[em.experiment_id] = exp_mol_counts.get(em.experiment_id, 0) + 1

    embed_images = len(molecules) <= IMAGE_THRESHOLD

    # ── Create workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # ── Sheet 1: Molecules ────────────────────────────────────────────────────
    ws_mol = wb.active
    ws_mol.title = "Molecules"

    mol_headers = [
        "Name", "SMILES", "Canonical SMILES", "Formula", "MW", "HBD", "HBA",
        "TPSA", "Rotatable Bonds", "InChIKey", "Method", "Date Created",
        "Notes", "Created By", "Structure",
    ]
    ws_mol.append(mol_headers)
    _style_header(ws_mol, len(mol_headers))

    if not embed_images:
        # Add a note in cell O1 (Structure header)
        ws_mol.cell(row=1, column=15).comment = None
        ws_mol.cell(row=1, column=15).value = "Structure (omitted)"
        note_row = [""] * 15
        note_row[0] = "Structure images omitted for large exports — see individual molecule pages"
        ws_mol.append(note_row)

    struct_col_letter = get_column_letter(15)  # column O
    ws_mol.column_dimensions[struct_col_letter].width = 14

    for i, mol in enumerate(molecules):
        row_num = i + 2 + (0 if embed_images else 1)  # +1 for note row
        ws_mol.append([
            mol.name,
            mol.smiles,
            mol.canonical_smiles or "",
            mol.molecular_formula or "",
            round(mol.molecular_weight, 4) if mol.molecular_weight else "",
            mol.hbd if mol.hbd is not None else "",
            mol.hba if mol.hba is not None else "",
            round(mol.tpsa, 4) if mol.tpsa is not None else "",
            mol.rotatable_bonds if mol.rotatable_bonds is not None else "",
            mol.inchikey or "",
            mol.method_used,
            mol.date_created.isoformat() if mol.date_created else "",
            mol.notes or "",
            emails.get(mol.created_by_user_id, "") if mol.created_by_user_id else "",
            "",  # Structure placeholder
        ])

        if embed_images and mol.svg_image:
            png_bytes = _svg_to_png(mol.svg_image)
            if png_bytes:
                img = XlImage(io.BytesIO(png_bytes))
                img.width = IMG_WIDTH_PX
                img.height = IMG_HEIGHT_PX
                ws_mol.add_image(img, f"{struct_col_letter}{row_num}")
                ws_mol.row_dimensions[row_num].height = ROW_HEIGHT_PT

    _auto_width(ws_mol, 14)  # skip Structure column

    # ── Sheet 2: Experiments ──────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Experiments")
    exp_headers = ["Title", "Date", "Notes", "Molecule Count", "Created By"]
    ws_exp.append(exp_headers)
    _style_header(ws_exp, len(exp_headers))

    for exp in experiments:
        ws_exp.append([
            exp.title,
            exp.date.isoformat() if exp.date else "",
            exp.notes or "",
            exp_mol_counts.get(exp.id, 0),
            emails.get(exp.created_by_user_id, "") if exp.created_by_user_id else "",
        ])

    _auto_width(ws_exp, len(exp_headers))

    # ── Sheet 3: Experiment-Molecule Links ────────────────────────────────────
    ws_links = wb.create_sheet("Experiment-Molecule Links")
    link_headers = ["Experiment Title", "Molecule Name", "Molecule SMILES"]
    ws_links.append(link_headers)
    _style_header(ws_links, len(link_headers))

    for em in exp_mol_rows:
        exp = exp_map.get(em.experiment_id)
        mol = mol_map.get(em.molecule_id)
        ws_links.append([
            exp.title if exp else "",
            mol.name if mol else "",
            mol.smiles if mol else "",
        ])

    _auto_width(ws_links, len(link_headers))

    # ── Sheet 4: Lab Info ─────────────────────────────────────────────────────
    ws_info = wb.create_sheet("Lab Info")
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 40

    info_rows = [
        ("Lab Name", lab.name),
        ("Lab Code", lab.lab_code),
        ("Created Date", lab.created_at.strftime("%Y-%m-%d") if lab.created_at else ""),
        ("Member Count", member_count),
        ("Export Date", date.today().isoformat()),
    ]
    for label, value in info_rows:
        ws_info.append([label, value])
        ws_info.cell(row=ws_info.max_row, column=1).font = HEADER_FONT

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
